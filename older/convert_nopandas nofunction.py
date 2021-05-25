from timeit import Timer
from itertools import islice
import openpyxl as px
from datetime import datetime as dt
from backports.zoneinfo import ZoneInfo as zi
t = Timer()

# Open the workbook and select a worksheet
wb = px.load_workbook('list_webex.xlsx')
sheet = wb['list_webex']
tzone_str = 'Europe/Brussels'

def parse_time(column, do):
    #merge date + hour into one string
    d = do['date_meeting']+' '+do[column+'_hour']+':00'
    #create a datetime object
    date = dt.strptime(d, '%d/%m/%Y %H:%M:%S').replace(tzinfo=zi(tzone_str))
    #make the date object timezone aware
    date.replace(tzinfo=zi(tzone_str))
    #return the date in isoformat
    return date.isoformat()

#list to hold header information
headers = []
#iterate through columns to fetch header and position
for i in sheet.iter_cols():
    headers.append(i[0].value)


# List to hold row values
mt_list = []
# Iterate through each row in worksheet and fetch values into tuples
for row in islice(sheet.values, 1, sheet.max_row):
    mt_list.append(row)


#Create a set of unique meetings
unimt_dicts = {}
for mt in mt_list:
    if not mt[0] in unimt_dicts:
        unimt_dict = {
        headers[0] : mt[0],
        headers[1] : mt[1],
        headers[2] : mt[2],
        headers[3] : mt[3]
        }
        unimt_dicts[mt[0]] = unimt_dict


inv_dicts = {}
for mt in unimt_dicts.items():
    inv_list = []
    inv_dict = {}
    #create a dict per row in the list, using the col header as key and col location as key and value index
    for i in mt_list:
        if mt[0] in i:
            inv_dict = {
            'displayname' : i[4]+' '+i[5],
            'email' : i[6]}
            inv_list.append(inv_dict.copy())     
    inv_dicts[mt[0]] = inv_list

# Create list to collect information to convert to json
treemt_dicts = {}
for mt in unimt_dicts.items():
    # replace mt by mt values to avoid unneeded '[1]' entries
    mt = mt[1]
    #put title in a var because it is used multiple times, readability
    title = mt['name_meeting']
    # Create dict object with mt name + aggregated members list
    treemt_dict = {
    'title' : title,
    'start' : parse_time('start', mt),
    'end' : parse_time('end', mt),
    'timezone' : tzone_str,
    'enabledAutoRecordmt' : False,
    'allowAnyUserToBeCoHost' : False,
    'invitees' : inv_dicts[title]
    }
    # Add one mt dict object to final dataset
    treemt_dicts[title] = treemt_dict

# 
print(treemt_dicts)
print(t.timeit())

print('a  ')
print(type(inv_list))
print('b  ')
print(type(inv_dict))
print('c  ')
print(type(inv_dicts))
print('d  ')
print(type(treemt_dict))
print('e  ')
print(type(treemt_dicts))
print('f  ')
print(type(unimt_dicts))
print('g ')
print(type(mt_list))
