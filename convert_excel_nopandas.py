from itertools import islice
import openpyxl as px
from datetime import datetime as dt
from backports.zoneinfo import ZoneInfo as zi

def parse_time(column, do, tzone_str):
    """**parse_time** - Function to parse strings to time zone aware 
    datetime objects, and then convert them to ISO8601 format strings

    :param column: name of the column to search value in
    :type column: string
    :param do: information about a unique meeting
    :type do: dict
    :param tzone_str: time zone
    :type tzone_str: string
    :return: iso8601 date
    :rtype: string
    """
    # merge date + hour into one string, choose method based on 
    # object type being string or datetime
    if type(do['date_meeting']) == dt:
        d_str = do['date_meeting']
        d = str(d_str.day)+'/'+str(d_str.month)+'/' + \
            str(d_str.year)+' '+do[column+'_hour'] + ':00'
    else:
        d = do['date_meeting']+' '+do[column+'_hour'] + ':00'
    date = dt.strptime(d, '%d/%m/%Y %H:%M:%S').replace(tzinfo=zi(tzone_str))
    date.replace(tzinfo=zi(tzone_str))
    return date.isoformat()


def get_headers(sheet):
    """get_headers Generate a list to be used as headers by
    reading in the first line of the worksheet.

    :param sheet: the required worksheet as a data object
    :type sheet: worksheet via openpyxl
    :return: list of header names
    :rtype: list
    """
    # list to hold header information
    headers = []
    # iterate through columns to fetch header and position
    for i in sheet.iter_cols():
        headers.append(i[0].value)
    return headers


def get_rows(sheet):
    """get_rows Generate a list populated with every row in
    the passed worksheet, skipping the header row.

    :param sheet: the required worksheet as a data object
    :type sheet: worksheet via openpyxl
    :return: all row entries
    :rtype: list
    """
    # List to hold row values
    mt_list = []
    # Iterate through each row in worksheet and fetch values into tuples
    for row in islice(sheet.values, 1, sheet.max_row):
        mt_list.append(row)
    return mt_list


def get_meetings(mt_list, headers):
    """get_meetings  Create a set of unique meetings based
    on the title. 

    :param mt_list: all worksheet row entries
    :type mt_list: list
    :param headers: list of header names
    :type headers: list
    :return: set of unique meetings
    :rtype: dict
    """

    unimt_dicts = {}
    for mt in mt_list:
        if not mt[0] in unimt_dicts:
            unimt_dict = {
                headers[0]: mt[0],
                headers[1]: mt[1],
                headers[2]: str(mt[2])[0:5],
                headers[3]: str(mt[3])[0:5]
            }
            unimt_dicts[mt[0]] = unimt_dict
    return unimt_dicts


def get_invitees(uni_mt, mt_list):
    """get_invitees Per meeting collect all attendees in one 
    list of dictionaries using the meeting name as key

    :param uni_mt: collection of unique meetings
    :type uni_mt: dict
    :param mt_list: all excel row entries
    :type mt_list: list
    :return: all meetings with their full invite list
    :rtype: dict
    """
    inv_dicts = {}
    for mt in uni_mt.items():
        inv_list = []
        inv_dict = {}
        for i in mt_list:
            if mt[0] in i:
                inv_dict = {
                    'displayname': i[4]+' '+i[5],
                    'email': i[6],
                    'coHost': i[7]}
                inv_list.append(inv_dict.copy())
        inv_dicts[mt[0]] = inv_list
    return inv_dicts


def main(file='list_webex.xlsx',tzone_str='Europe/Brussels'):
    """**main** - Converts an excel file to a dict suitable for creating
    webex meetings via the Webex REST api with lists and dicts.

    :param file: source data do convert, defaults to 'list_webex.xlsx'
    :type file: str, optional
    :param tzone_str: desired timezone, defaults to 'Europe/Brussels'
    :type tzone_str: str, optional
    """

    file = 'list_webex.xlsx'
    wb = px.load_workbook(file)
    sheet = wb['list_webex']
    headers = get_headers(sheet)
    rows = get_rows(sheet)
    mt_dict = get_meetings(rows, headers)
    invitees = get_invitees(mt_dict, rows)
    mts_tree = {}

    for mt in mt_dict.items():
        mt = mt[1]
        title = mt['name_meeting']
        mt_tree = {
            'title': title,
            'start': parse_time('start', mt, tzone_str),
            'end': parse_time('end', mt, tzone_str),
            'timezone': tzone_str,
            'enabledAutoRecordMeeting': False,
            'allowAnyUserToBeCoHost': False,
            'invitees': invitees[title]
        }
        # Add one mt dict object to final dataset
        mts_tree[title] = mt_tree
    return mts_tree


# execute main when called directly
if __name__ == '__main__':
    main()
