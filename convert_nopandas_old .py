import json
from collections import OrderedDict
import numpy as np
from itertools import islice
from openpyxl import load_workbook
import datetime as dt

# Open the workbook and select a worksheet
wb = load_workbook('list_webex.xlsx')
sheet = wb['list_webex']

# List to hold dictionaries
meeting_list = {}
# Iterate through each row in worksheet and fetch values into dict
for row in islice(sheet.values, 0, sheet.max_row):
    meeting_list.append(row)


#Create a set of unique meeting
unimeeting_list = []
for meeting in meeting_list:
    if not unimeeting_list.__contains__(meeting[0]):
        unimeeting_list.append(meeting[0])

# Create list to collect information to convert to json
treemeeting_list = []
for meeting in unimeeting_list:
    members_list = []
    # Run through participants list and create a dict object per member
    for item in meeting_list:
        if meeting == item[0]:
            print(item)
            member = OrderedDict()
            member['displayname'] = item['sn_participant']+' '+item['sn_participant']
            member['email'] = item['email_participant']
            # Add dict object to members list
            members_list.append(member)
    # Create dict object with meeting name + aggregated members list
    treemeeting_dict = OrderedDict()
    treemeeting_dict['meeting_name'] = meeting
    treemeeting_dict['meeting_members'] = members_list
    # Add one meeting dict object to final dataset
    treemeeting_list.append(treemeeting_dict)

# Serialize the list of dicts to JSON
j = json.dumps(treemeeting_list)

print(j)


