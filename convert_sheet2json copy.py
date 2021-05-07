import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook

# Open the workbook and select a worksheet
wb = load_workbook('list_webex.xlsx')
sheet = wb['list_webex']
# List to hold dictionaries
rooms_list = []
# Iterate through each row in worksheet and fetch values into dict
for row in islice(sheet.values, 1, sheet.max_row):
    rooms_list.append(row)

#Create a set of unique rooms
unirooms_list = []
for room in rooms_list:
    if not unirooms_list.__contains__(room[0]):
        unirooms_list.append(room[0])

# Create list to collect information to convert to json
treerooms_list = []
for room in unirooms_list:
    members_list = []
    # Run through participants list and create a dict object per member
    for item in rooms_list:
        if room == item[0]:
            member = OrderedDict()
            member['room_breakout'] = item[1]
            member['participant_fn'] = item[2]
            member['participant_sn'] = item[3]
            member['participant_mail'] = item[4]
            # Add dict object to members list
            members_list.append(member)
    # Create dict object with room name + aggregated members list
    treeroom_dict = OrderedDict()
    treeroom_dict['room_name'] = room
    treeroom_dict['room_members'] = members_list
    # Add one room dict object to final dataset
    treerooms_list.append(treeroom_dict)

# Serialize the list of dicts to JSON
j = json.dumps(treerooms_list)

print(j)